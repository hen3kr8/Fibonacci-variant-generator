import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
from ScaledData import scale_data
from BuildModels import create_tensors
from keras.optimizers import Adam
from keras.losses import MeanSquaredError
import os
from tensorflow.python.ops.numpy_ops import np_config
from openpyxl import Workbook
from openpyxl import load_workbook
from typing import Tuple, List

np_config.enable_numpy_behavior()

X_TIMESERIES_SCALER_PATH = 'ScaledData/Scalers/x_timeseries_scaler.save'
Y_TIMESERIES_SCALER_PATH = 'ScaledData/Scalers/y_timeseries_scaler.save'

SIMPLE_X_TIMESERIES_SCALER_PATH = 'ScaledData/Scalers/simple_x_timeseries_scaler.save'
SIMPLE_Y_TIMESERIES_SCALER_PATH = 'ScaledData/Scalers/simple_y_timeseries_scaler.save'

SHOW_TEST_RESULTS = False
BATCH_SIZE = 1
RETURN_SEQUENCES = True
USE_SIMPLE_FIB_DATA = True

def save_model(model, model_name):
    model.save(f'{model_name}.h5')

def write_run_to_excel(results_for_excel:dict):
    wb = None
    ws = None
    try:
        if not os.path.exists(results_for_excel['excel_file']):
            wb = Workbook()
            ws = wb.active
            # write header
            training_acc_header = ['training_accuracy_ep_' + str(i) for i in range(results_for_excel['epochs'])]
            val_acc_header = ['val_accuracy_ep_' + str(i) for i in range(results_for_excel['epochs'])]

            file_headers = list(results_for_excel.keys())[:-2]
            file_headers.extend(training_acc_header)
            file_headers.extend(val_acc_header)
            ws.append(file_headers)

        else:
            wb = load_workbook(filename=results_for_excel['excel_file'])
            ws = wb.active
        ws.append([results_for_excel['excel_file'], 
                results_for_excel['run'], 
                results_for_excel['learning_rate'], 
                results_for_excel['epochs'], 
                results_for_excel['num_layers'],
                results_for_excel['num_units'],
                results_for_excel['test_mse'],
                results_for_excel['final_training_mse'], 
                *results_for_excel['mse'].tolist(), 
                *results_for_excel['val_mse'].tolist(),
        ]) 
        wb.save(results_for_excel['excel_file'])
    except Exception as e:
        print(e)
        raise e

# convert data back from tensor to numpy array and scale back to original scale
def convert_tensor_to_numpy_and_scale_back(X_data, y_data, simple=False):
    if simple:
        x_timeseries_scaler = scale_data.load_scaler(SIMPLE_X_TIMESERIES_SCALER_PATH)
        y_timeseries_scaler = scale_data.load_scaler(SIMPLE_Y_TIMESERIES_SCALER_PATH)
    else:
        x_timeseries_scaler = scale_data.load_scaler(X_TIMESERIES_SCALER_PATH)
        y_timeseries_scaler = scale_data.load_scaler(Y_TIMESERIES_SCALER_PATH)

    X_reshaped = X_data
    if len(X_data.shape) == 3:
        X_reshaped = tf.squeeze(X_reshaped, axis=-1) 
    if len(y_data.shape) == 3:
        y_data = tf.squeeze(y_data, axis=-1) 

    x_timeseries_normal = x_timeseries_scaler.inverse_transform(X_reshaped)
    y_normal = y_timeseries_scaler.inverse_transform(y_data)
    return x_timeseries_normal, y_normal


def test_single_sample_static_and_timeseries(
    model: tf.keras.Model, 
    X: Tuple[tf.Tensor, tf.Tensor], 
    y: np.ndarray
    ):

    X_static = X[0]
    X_timeseries = X[1]
    X_static = X_static.reshape(1, X_static.shape[0])
    X_timeseries = X_timeseries[None, :, :]
    y_pred = model.predict(x = [X_static, X_timeseries], verbose=1, batch_size=BATCH_SIZE)
    x_timeseries_normal, y_test_normal = convert_tensor_to_numpy_and_scale_back(X_timeseries, y, simple=False)
    x_timeseries_normal, y_pred_normal = convert_tensor_to_numpy_and_scale_back(X_timeseries, y_pred, simple=False)
    
    if SHOW_TEST_RESULTS:        
        print("x: ", x_timeseries_normal)
        print("y_pred: ", y_pred_normal)
        print("y_test: ", y_test_normal)

def test_model(
        model: tf.keras.Model, 
        X_data: Tuple[tf.Tensor, tf.Tensor], 
        y_test:tf.Tensor, 
        static=False,
        x_scaler_path=None, 
        y_scaler_path=None, 
        scale=False
        ) -> List[float]:
    '''
    Evaluates the model on test data
    if return_sequences = True, then the RNN predicts a value for every time step.
    This makes the model more accurate
    else it only predicts the last value in the sequence.

    Parameters:
        model: The trained model
        X_data: The test data [X_static, X_timeseries]
        y_test: The test target variable
        batch_size: The batch size
        static: If the model has static data
        x_scaler_path: The path to the x scaler
        y_scaler_path: The path to the y scaler
        scale: If the data should be scaled back to original scale
    '''

    x_timeseries_scaler = None
    y_timeseries_scaler = None
    if x_scaler_path and y_scaler_path:
        if not os.path.exists(x_scaler_path):
            raise FileNotFoundError(f'{x_scaler_path} not found')
        
        x_timeseries_scaler = scale_data.load_scaler(x_scaler_path)
        y_timeseries_scaler = scale_data.load_scaler(y_scaler_path)

    else:
        x_scaler_path = X_TIMESERIES_SCALER_PATH
        y_scaler_path = Y_TIMESERIES_SCALER_PATH
        x_timeseries_scaler = scale_data.load_scaler(x_scaler_path)
        y_timeseries_scaler = scale_data.load_scaler(y_scaler_path)

    results = None
    if static:
        results = model.evaluate(
            x=[X_data[0], X_data[1]], y=y_test, 
            verbose=1, batch_size=BATCH_SIZE
        )
    else:
        results = model.evaluate(
            X_data, y_test, 
            verbose=1, batch_size=BATCH_SIZE
            )

    if SHOW_TEST_RESULTS:
        print('test results', results)
        print("Test loss:", results[0])
        print("Test accuracy:", results[1])
    
    y_pred = None
    if static:
        y_pred = model.predict(
            x =[X_data[0], X_data[1]], 
            verbose=1, 
            batch_size=BATCH_SIZE
        )
    else:
        y_pred = model.predict(x =X_data, verbose=1, batch_size=BATCH_SIZE)

    # convert x and y back to original scale
    X_reshaped = None
    if not static:
        X_reshaped = X_data.reshape(-1, X_data.shape[2])  # Reshape to 2D
    else:
        X_reshaped = X_data[1].reshape(-1, X_data[1].shape[2])
    Y_test_reshaped = y_test.reshape(-1, 1)  # Reshape to 2D
    Y_pred_reshaped = y_pred.reshape(-1, 1)  # Reshape to 2D

    x_timeseries_normal = x_timeseries_scaler.inverse_transform(X_reshaped)
    y_test_normal = y_timeseries_scaler.inverse_transform(Y_test_reshaped)
    y_pred_normal = y_timeseries_scaler.inverse_transform(Y_pred_reshaped)

    if SHOW_TEST_RESULTS:
        print("x: ", x_timeseries_normal) 
        print("y_pred: ", y_pred_normal)
        print("y_test: ", y_test_normal)

    return results
